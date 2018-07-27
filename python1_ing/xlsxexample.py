def readFromFile(filename):
    import openpyxl
    from openpyxl import load_workbook,Workbook
    wb = load_workbook(filename)
    ws = wb.active
    wiersze = []
    kolumny=['name','description']
    for y in range(2,ws.max_row+1):
        name=ws.cell(row=y,column=1).value
        description=ws.cell(row=y,column=2).value
        wiersze.append({'name':name,'description':description})
    #read 
    return wiersze
	
def writeToFile(filename,tabelka,kolumny):
    import openpyxl
    from openpyxl import load_workbook,Workbook
    wb = Workbook()
    for i in range(len(tabelka)):
	    ws.cell(row=2+i,column=kolumny.index('name')+1).value=tabelka[i]['name']
		ws.cell(row=2+i,column=kolumny.index('description')+1).value=tabelka[i]['description']
    wb.save(filename)

if __name__=='__main__':
    t=readFromFile('nazwapliku.xlsx')
